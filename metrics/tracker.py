"""Metrics Tracker — Excel-based benchmark persistence.

Reads and writes run metrics to benchmark_data.xlsx using openpyxl.
The file is auto-created with styled headers on first use.
"""

import os
from datetime import datetime
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Config ────────────────────────────────────────────────────────────────────

METRICS_FILE = os.path.join(os.path.dirname(__file__), "benchmark_data.xlsx")

COLUMNS = [
    "Run ID",
    "Timestamp",
    "Pipeline",
    "Prompt (truncated)",
    "Test Pass Rate (%)",
    "Tests Passed",
    "Tests Failed",
    "Avg Lint Score (/10)",
    "Security Score (/10)",
    "Fix Loops",
    "Lines of Code",
    "Time (seconds)",
    "PR Mock File",
]

COL_WIDTHS = [12, 22, 14, 45, 18, 14, 14, 20, 20, 12, 15, 15, 40]

# Header style: purple background, white bold text
_HEADER_FILL = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Alternating row fill
_ODD_FILL  = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
_EVEN_FILL = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
_ROW_FONT  = Font(color="E2E8F0", size=10)
_CENTER    = Alignment(horizontal="center")

_THIN_BORDER = Border(
    left=Side(style="thin", color="3F3F6E"),
    right=Side(style="thin", color="3F3F6E"),
    top=Side(style="thin", color="3F3F6E"),
    bottom=Side(style="thin", color="3F3F6E"),
)


# ── Internal helpers ──────────────────────────────────────────────────────────

def _bootstrap_workbook() -> openpyxl.Workbook:
    """Create a new workbook with styled headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchmark Data"
    ws.row_dimensions[1].height = 36

    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border    = _THIN_BORDER
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width

    # Freeze header row
    ws.freeze_panes = "A2"
    wb.save(METRICS_FILE)
    return wb


def _load_wb():
    """Load existing workbook or create it fresh."""
    if os.path.exists(METRICS_FILE):
        return openpyxl.load_workbook(METRICS_FILE)
    return _bootstrap_workbook()


def _style_data_row(ws, row_idx: int, num_cols: int) -> None:
    fill = _ODD_FILL if row_idx % 2 == 1 else _EVEN_FILL
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill      = fill
        cell.font      = _ROW_FONT
        cell.alignment = _CENTER
        cell.border    = _THIN_BORDER


# ── Public API ────────────────────────────────────────────────────────────────

def save_run_metrics(state: dict, pipeline_type: str) -> None:
    """
    Append one row of metrics for the given run to benchmark_data.xlsx.

    Parameters
    ----------
    state         : dict   Final pipeline state.
    pipeline_type : str    "Multi-Agent" or "Baseline".
    """
    wb = _load_wb()
    ws = wb.active

    generated_code = state.get("generated_code", {})
    lint_scores    = state.get("lint_scores", {})
    test_results   = state.get("test_results", {})
    run_dir        = state.get("run_dir", "")
    requirement    = state.get("requirement", "")

    total_lines = sum(len(c.splitlines()) for c in generated_code.values())
    avg_lint = (
        round(sum(lint_scores.values()) / len(lint_scores), 1) if lint_scores else 0.0
    )
    pr_path = os.path.join(run_dir, "github_pr_mock.json") if run_dir else "N/A"
    prompt_preview = (requirement[:42] + "…") if len(requirement) > 42 else requirement

    row_data = [
        state.get("run_id", "N/A"),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        pipeline_type,
        prompt_preview,
        test_results.get("pass_rate", 0.0),
        test_results.get("passed_count", 0),
        test_results.get("failed_count", 0),
        avg_lint,
        state.get("security_score", 0.0),
        state.get("fix_attempts", 0),
        total_lines,
        state.get("elapsed_time", 0.0),
        pr_path,
    ]

    next_row = ws.max_row + 1
    ws.append(row_data)
    _style_data_row(ws, next_row, len(COLUMNS))

    wb.save(METRICS_FILE)


def load_all_metrics() -> List[Dict[str, Any]]:
    """
    Load every data row from benchmark_data.xlsx.

    Returns
    -------
    list of dicts — one dict per run, keyed by column name.
    """
    if not os.path.exists(METRICS_FILE):
        return []

    wb  = openpyxl.load_workbook(METRICS_FILE, read_only=True, data_only=True)
    ws  = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(dict(zip(COLUMNS, row)))
    wb.close()
    return rows


def get_metrics_file_path() -> str:
    """Return the absolute path to the Excel metrics file."""
    return METRICS_FILE
